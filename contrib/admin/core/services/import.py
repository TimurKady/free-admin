# -*- coding: utf-8 -*-
"""
import

Import service with CSV/JSON/XLSX support and reporting.

Version: 0.1.0
Author: Timur Kady
Email: timurkady@yandex.com
"""

from __future__ import annotations

import asyncio
import csv
import json
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from itertools import islice
from pathlib import Path
from typing import Any, Iterable, Sequence
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import load_workbook

from ..model import ModelAdmin


@dataclass
class ImportReport:
    processed: int = 0
    created: int = 0
    updated: int = 0
    errors: list[str] = field(default_factory=list)


@dataclass
class CachedUpload:
    path: Path
    fmt: str
    expires_at: datetime


class BaseParser:
    """Base parser returning row dictionaries from a file path."""

    def parse(self, path: Path) -> Iterable[dict[str, Any]]:
        raise NotImplementedError


class ParserRegistry:
    """Registry mapping file formats to parser classes."""

    def __init__(self) -> None:
        self._parsers: dict[str, type[BaseParser]] = {}

    def register(self, *fmts: str):
        def decorator(cls: type[BaseParser]):
            for fmt in fmts:
                self._parsers[fmt] = cls
            return cls
        return decorator

    def get(self, fmt: str) -> BaseParser:
        parser_cls = self._parsers.get(fmt.lower())
        if not parser_cls:
            raise ValueError(f"Unsupported format: {fmt}")
        return parser_cls()


parser_registry = ParserRegistry()


@parser_registry.register("csv")
class CsvParser(BaseParser):
    """Yield dictionaries from CSV files."""

    def parse(self, path: Path) -> Iterable[dict[str, Any]]:
        with path.open("r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield row


@parser_registry.register("json")
class JsonParser(BaseParser):
    """Yield dictionaries from JSON arrays."""

    def parse(self, path: Path) -> Iterable[dict[str, Any]]:
        decoder = json.JSONDecoder()
        buffer = ""
        with path.open("r", encoding="utf-8") as f:
            chunk = f.read(4096)
            while chunk:
                buffer += chunk
                while True:
                    buffer = buffer.lstrip()
                    if buffer.startswith("["):
                        buffer = buffer[1:]
                    if buffer.startswith("]"):
                        return
                    try:
                        obj, index = decoder.raw_decode(buffer)
                    except ValueError:
                        break
                    yield obj
                    buffer = buffer[index:]
                    buffer = buffer.lstrip()
                    if buffer.startswith(","):
                        buffer = buffer[1:]
                chunk = f.read(4096)


@parser_registry.register("xlsx", "xls")
class XlsxParser(BaseParser):
    """Yield dictionaries from XLSX spreadsheets."""

    def parse(self, path: Path) -> Iterable[dict[str, Any]]:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers: list[str] = list(next(rows, []))
        for values in rows:
            yield {str(h): v for h, v in zip(headers, values)}


class UploadStep:
    """Retrieve cached upload ensuring it exists and is fresh."""

    def __init__(self, cache: dict[str, CachedUpload], cleanup_cb) -> None:
        self._cache = cache
        self._cleanup = cleanup_cb

    def run(self, token: str) -> CachedUpload:
        info = self._cache.get(token)
        if not info or info.expires_at < datetime.now():
            self._cleanup(token)
            raise HTTPException(status_code=404)
        return info


class ParsingStep:
    """Yield rows from cached file using registered parsers."""

    def __init__(self, registry: ParserRegistry = parser_registry) -> None:
        self.registry = registry

    def run(self, info: CachedUpload) -> Iterable[dict[str, Any]]:
        parser = self.registry.get(info.fmt)
        return parser.parse(info.path)


class FieldFilterStep:
    """Limit row fields to selected ones."""

    def run(
        self, rows: Iterable[dict[str, Any]], fields: Sequence[str]
    ) -> Iterable[dict[str, Any]]:
        allowed = list(dict.fromkeys(fields))
        for row in rows:
            yield {k: row.get(k) for k in allowed}


class PersistenceStep:
    """Create or update records via ``ModelAdmin``."""

    async def run(
        self,
        rows: Iterable[dict[str, Any]],
        admin: ModelAdmin,
        report: ImportReport,
        dry: bool,
    ) -> None:
        for row in rows:
            report.processed += 1
            if dry:
                continue
            obj, created = await admin.get_or_create_for_import(row)
            if created:
                report.created += 1
            else:
                await admin.update_for_import(obj, row)
                report.updated += 1


class ImportPipeline:
    """Coordinate sequential import steps."""

    def __init__(
        self,
        upload_step: UploadStep,
        parsing_step: ParsingStep,
        field_step: FieldFilterStep,
        persistence_step: PersistenceStep,
    ) -> None:
        self.upload_step = upload_step
        self.parsing_step = parsing_step
        self.field_step = field_step
        self.persistence_step = persistence_step

    async def run(
        self,
        token: str,
        admin: ModelAdmin,
        fields: Sequence[str],
        dry: bool,
    ) -> ImportReport:
        info = self.upload_step.run(token)
        rows = self.parsing_step.run(info)
        filtered = self.field_step.run(rows, fields)
        report = ImportReport()
        await self.persistence_step.run(filtered, admin, report, dry)
        return report


class ImportService:
    """Handle import workflow with temporary file caching."""

    def __init__(
        self,
        tmp_dir: str | None = None,
        ttl: int = 300,
        max_size: int = 10 * 1024 * 1024,
        registry: ParserRegistry = parser_registry,
    ) -> None:
        self.tmp_dir = tmp_dir or tempfile.gettempdir()
        self.ttl = ttl
        self.max_size = max_size
        self._cache: dict[str, CachedUpload] = {}
        self._registry = registry

    async def import_rows(
        self, admin: ModelAdmin, rows: Iterable[dict[str, Any]]
    ) -> int:
        """Create or update objects for ``admin`` from ``rows``."""
        count = 0
        for row in rows:
            obj, created = await admin.get_or_create_for_import(row)
            if not created:
                await admin.update_for_import(obj, row)
            count += 1
        return count

    async def cache_upload(self, upload) -> str:
        """Store ``upload`` in a temporary file and return its token."""
        suffix = Path(getattr(upload, "filename", "")).suffix or ".dat"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir=self.tmp_dir)
        if hasattr(upload, "file"):
            upload.file.seek(0)
            await asyncio.to_thread(shutil.copyfileobj, upload.file, tmp)
        else:
            await asyncio.to_thread(tmp.write, upload)
        tmp.close()
        path = Path(tmp.name)
        if path.stat().st_size > self.max_size:
            await asyncio.to_thread(path.unlink)
            raise HTTPException(status_code=413, detail="File too large")
        token = uuid4().hex
        fmt = suffix.lstrip(".").lower()
        expires_at = datetime.now() + timedelta(seconds=self.ttl)
        self._cache[token] = CachedUpload(path, fmt, expires_at)
        asyncio.get_running_loop().call_later(
            self.ttl, lambda: self.cleanup(token)
        )
        return token

    async def preview(
        self, token: str, fields: Sequence[str], limit: int = 20
    ) -> list[dict[str, Any]]:
        """Return first rows from cached file for preview."""
        upload_step = UploadStep(self._cache, self.cleanup)
        parsing_step = ParsingStep(self._registry)
        field_step = FieldFilterStep()
        info = upload_step.run(token)
        rows = parsing_step.run(info)
        filtered = field_step.run(rows, fields)
        return list(islice(filtered, limit))

    async def run(
        self,
        admin: ModelAdmin,
        token: str,
        fields: Sequence[str] | None = None,
        dry: bool = False,
    ) -> ImportReport:
        """Import cached file for ``admin`` returning an ``ImportReport``."""
        pipeline = ImportPipeline(
            UploadStep(self._cache, self.cleanup),
            ParsingStep(self._registry),
            FieldFilterStep(),
            PersistenceStep(),
        )
        fields = list(dict.fromkeys(fields or admin.get_import_fields()))
        return await pipeline.run(token, admin, fields, dry)

    def cleanup(self, token: str) -> None:
        info = self._cache.pop(token, None)
        if info:
            loop = asyncio.get_running_loop()
            loop.create_task(asyncio.to_thread(info.path.unlink))

# The End

